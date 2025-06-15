import os
from openpyxl import Workbook, load_workbook
from django.conf import settings

def log_user_activity_to_excel(email, activity_type):
    filename = os.path.join(settings.MEDIA_ROOT, "user_activity.xlsx")

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity Log"
        ws.append(["Email", "Activity", "Timestamp"])
    else:
        wb = load_workbook(filename)
        ws = wb.active

    from datetime import datetime
    ws.append([email, activity_type, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(filename)
